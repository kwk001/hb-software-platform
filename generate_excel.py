#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = '页面开发任务列表'

# 定义表头
headers = [
    '任务名称', '项目名称', '任务类型', '指派给', '研发人员', '任务状态', 
    '期望完成时间', '计划开始时间', '计划结束时间', '优先级', '预估工作量（天）',
    '当前进度', '剩余工作量（天）', '实际工作量（天）', '实际开始时间', '实际结束时间',
    '是否合入产品', '需求是否MD文档', '需求详细程度（数值1-10）', '上级任务名称',
    '任务内容', '原因分析及处理方法', 'AI风险识别', '图片上传', '附件',
    '上级任务编码', 'AI风险等级', 'AI风险说明', 'AI建议工作量(天)', 'AI分析时间',
    '提交人', '创建时间', '更新时间', '更新人'
]

# 设置表头
ws.append(headers)

# 设置表头样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 页面数据
pages = [
    ('门户首页', '门户端', '/', '平台门户首页展示，包含轮播、统计、功能介绍、软件展示、政策展示'),
    ('软件产品列表', '门户端', '/software', '软件产品列表展示，支持分类筛选和搜索'),
    ('软件产品详情', '门户端', '/software/:id', '软件产品详细信息展示'),
    ('政策中心', '门户端', '/policy', '政策列表展示和查看'),
    ('需求广场', '门户端', '/demand', '需求对接展示'),
    ('用户登录', '门户端', '/login', '用户登录（账号密码/手机号验证码）'),
    ('用户注册', '门户端', '/register', '用户注册（手机号+验证码+密码）'),
    ('企业中心首页', '企业端', '/enterprise', '企业端Dashboard，展示统计信息和快捷入口'),
    ('企业入驻申请', '企业端', '/enterprise/apply', '企业入驻申请表单（基本信息+资质信息）'),
    ('软件发布', '企业端', '/enterprise/software/publish', '软件产品发布（基本信息+功能介绍+附件上传）'),
    ('我的软件', '企业端', '/enterprise/software/list', '已发布软件列表管理'),
    ('补贴申报', '企业端', '/enterprise/subsidy/apply', '补贴券申报（选择软件+填写信息+上传附件）'),
    ('我的补贴', '企业端', '/enterprise/subsidy/list', '补贴申报记录和使用管理'),
    ('需求提交', '企业端', '/enterprise/demands', '需求发布和管理'),
    ('留言提交', '企业端', '/enterprise/messages', '向平台提交留言（个性化需求/咨询/建议）'),
    ('消息中心', '企业端', '/enterprise/message-center', '站内消息查看和管理'),
    ('通知设置', '企业端', '/enterprise/notification-settings', '消息通知方式设置'),
    ('平台首页', '平台端', '/platform', '平台端Dashboard，展示运营统计数据'),
    ('企业审核', '平台端', '/platform/audit/enterprise', '企业入驻申请审核管理'),
    ('软件审核', '平台端', '/platform/audit/software', '软件产品发布审核管理'),
    ('补贴审核', '平台端', '/platform/audit/subsidy', '补贴券申报审核管理'),
    ('需求汇总', '平台端', '/platform/demands', '需求对接汇总和管理'),
    ('数据统计', '平台端', '/platform/statistics', '运营数据统计和多维度分析'),
    ('政策管理', '平台端', '/platform/policy', '政策发布、编辑、下架、置顶管理'),
    ('用户管理', '平台端', '/platform/users', '平台用户管理'),
    ('角色管理', '平台端', '/platform/role', '角色权限配置管理'),
    ('菜单管理', '平台端', '/platform/menu', '系统菜单配置管理'),
    ('部门管理', '平台端', '/platform/dept', '部门组织架构管理'),
    ('岗位管理', '平台端', '/platform/post', '岗位职位管理'),
    ('留言管理', '平台端', '/platform/messages', '企业留言查看和回复'),
    ('数据字典', '平台端', '/platform/dict', '数据字典管理（多级字典）'),
    ('操作日志', '平台端', '/platform/logs/operation', '系统操作日志查询'),
    ('登录日志', '平台端', '/platform/logs/login', '用户登录日志查询'),
    ('消息模板', '平台端', '/platform/message-template', '消息通知模板配置'),
    ('消息发送', '平台端', '/platform/message-send', '站内消息发送管理'),
    ('发送记录', '平台端', '/platform/message-record', '消息发送记录查询'),
]

# 填充数据
now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
for idx, (task_name, module, path, desc) in enumerate(pages, 1):
    row = [
        task_name,  # 任务名称
        '湖北省工业软件公共服务平台',  # 项目名称
        '页面开发',  # 任务类型
        '',  # 指派给
        '',  # 研发人员
        '已完成',  # 任务状态
        '2026-03-20',  # 期望完成时间
        '2026-03-01',  # 计划开始时间
        '2026-03-17',  # 计划结束时间
        '高' if idx <= 29 else '中',  # 优先级
        2,  # 预估工作量（天）
        '100%',  # 当前进度
        0,  # 剩余工作量（天）
        2,  # 实际工作量（天）
        '2026-03-01',  # 实际开始时间
        '2026-03-17',  # 实际结束时间
        '是',  # 是否合入产品
        '是',  # 需求是否MD文档
        8,  # 需求详细程度（数值1-10）
        module,  # 上级任务名称
        f'页面路径: {path}\n功能描述: {desc}',  # 任务内容
        '',  # 原因分析及处理方法
        '低',  # AI风险识别
        '',  # 图片上传
        '',  # 附件
        f'TASK-{idx:03d}',  # 上级任务编码
        '低',  # AI风险等级
        '标准页面开发，技术成熟，风险可控',  # AI风险说明
        2,  # AI建议工作量(天)
        now,  # AI分析时间
        'AI助手',  # 提交人
        now,  # 创建时间
        now,  # 更新时间
        'AI助手',  # 更新人
    ]
    ws.append(row)

# 设置列宽
column_widths = [15, 25, 12, 10, 10, 10, 12, 12, 12, 8, 12, 10, 12, 12, 12, 12, 12, 12, 15, 12, 40, 20, 10, 10, 10, 12, 10, 30, 12, 18, 10, 18, 18, 10]
for idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(idx)].width = width

# 设置数据行样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

# 冻结首行
ws.freeze_panes = 'A2'

# 保存文件
output_file = '页面开发任务列表_完整模版.xlsx'
wb.save(output_file)
print(f'Excel文件已生成: {output_file}')
print(f'包含 {len(pages)} 个页面任务')
